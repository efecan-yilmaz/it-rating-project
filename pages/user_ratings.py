import streamlit as st
from utils.utils import JSON_DETAILS_DATA_PATH, load_details_data_from_json, TEMPLATE_PATH, JSON_USER_RATINGS_PATH
import pandas as pd
import openpyxl
from io import BytesIO
import json
import os
import uuid

from utils.process_locator import determine_page, save_current_page, Page, run_redirect, clean_for_previous_direction

run_redirect(Page.USER_RATINGS.value)

st.title("User Ratings Collection")

details_df = load_details_data_from_json(JSON_DETAILS_DATA_PATH)
valid_ids = set(details_df["id"].astype(str))

# Read existing ratings from JSON on initialization
if os.path.exists(JSON_USER_RATINGS_PATH):
    with open(JSON_USER_RATINGS_PATH, "r") as f:
        all_data = json.load(f)
    
    df_display = pd.DataFrame(all_data).drop(columns=["id", "details_id"])
    st.dataframe(df_display)
else:
    all_data = []

if details_df.empty:
    st.info("No tools added yet. Please follow previous steps to add tools to collect user ratings.")
else:
    # Load template
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # Fill "tool" column starting from 5A
    for idx, tool in enumerate(details_df["tool"], start=5):
        ws[f"A{idx}"] = tool

    # Fill "category" column starting from 5B
    for idx, category in enumerate(details_df["category"], start=5):
        ws[f"B{idx}"] = category

    # Fill "id" column starting from 5J
    for idx, id_val in enumerate(details_df["id"], start=5):
        ws[f"J{idx}"] = id_val

    # Hide the "id" column (column J)
    ws.column_dimensions['J'].hidden = True

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.download_button(
        label="⬇️ Download User Rating Collection Excel",
        data=output,
        file_name="user_rating_collection.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.write("Upload your user ratings in the provided Excel format. Ensure that the ratings are between 1 and 5 for each category.")
    st.write("You can upload multiple files at once. The uploaded data will be merged with existing ratings.")

    if "file_uploader_key" not in st.session_state:
        st.session_state["file_uploader_key"] = 0

    uploaded_files = st.file_uploader(
        "📤 Upload User Rating Collection Excel",
        type=["xlsx"],
        accept_multiple_files=True,
        key=st.session_state["file_uploader_key"]
    )

    new_data = []
    if uploaded_files:
        for uploaded_file in uploaded_files:
            df = pd.read_excel(uploaded_file, header=None)
            for idx in range(4, len(df)):
                row = df.iloc[idx]
                numeric_cols = row[2:9]
                details_id = row[9] if len(row) > 9 and pd.notna(row[9]) else ""
                # Check if details_id exists in valid_ids
                if str(details_id) not in valid_ids:
                    continue
                if all(
                    pd.api.types.is_number(x) and 1 <= x <= 5
                    for x in numeric_cols
                ):
                    tool = row[0] if pd.notna(row[0]) and str(row[0]).strip() else "Manual Task"
                    new_data.append({
                        "id": str(uuid.uuid4()),
                        "Tool/Task": tool,
                        "Category": row[1],
                        "Frequency of Use": row[2],
                        "Time Efficiency": row[3],
                        "Output Quality": row[4],
                        "Ease of Use": row[5],
                        "Integration": row[6],
                        "Reliability": row[7],
                        "Satisfaction": row[8],
                        "details_id": details_id
                    })
        if new_data:
            all_data.extend(new_data)
            # Save updated data to JSON
            with open(JSON_USER_RATINGS_PATH, "w") as f:
                json.dump(all_data, f, indent=2)
                st.session_state["file_uploader_key"] += 1
                st.rerun()

if st.button("🗑️ Delete All User Ratings"):
    if os.path.exists(JSON_USER_RATINGS_PATH):
        os.remove(JSON_USER_RATINGS_PATH)
        st.success("All user ratings have been deleted.")
        st.rerun()
    else:
        st.warning("No ratings file found to delete.")

next_step_enabled = os.path.exists(JSON_USER_RATINGS_PATH)
if not next_step_enabled:
    st.warning("Please complete User Ratings Collection before proceeding to the next step.")

col_prev_next = st.columns([0.5, 0.5])
with col_prev_next[0]:
    if st.button("⬅️ Previous step"):
        save_current_page(Page.DETAIL_DATA)
        clean_for_previous_direction(Page.USER_RATINGS)
        st.switch_page(Page.DETAIL_DATA.value)
with col_prev_next[1]:
    if st.button("➡️ Next step", disabled=not next_step_enabled):
        save_current_page(Page.REQUIREMENT_ENGINEERING)
        st.switch_page(Page.REQUIREMENT_ENGINEERING.value)